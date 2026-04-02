"""
Seafile Grade Monitor
=====================
Run this script whenever you want to check for new grades.
- First run: silently records all existing files (no notifications).
- Every run after: notifies you only about NEW files uploaded since last time.

Requirements:
    pip install requests openpyxl pdfplumber plyer
"""

import os
import re
import sys
import csv
import json
import hashlib
import logging
import requests
import pdfplumber
import openpyxl
from io import BytesIO
from pathlib import Path
from datetime import datetime

# ─────────────────────────────────────────────
#  CONFIG — edit everything in this section
# ─────────────────────────────────────────────
CONFIG = {
    "seafile_url":    "https://seafile.unistra.fr",
    "username":       "",
    "password":       "",
    "library_name":   "",
    "results_path":   "",
    "student_number": "",
    "state_file":     "seafile_monitor_state.json",
}
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


class SeafileClient:
    def __init__(self, url, username, password):
        self.base = url.rstrip("/")
        self.session = requests.Session()
        self._login(username, password)

    def _login(self, username, password):
        log.info("Logging in to Seafile ...")
        r = self.session.post(
            f"{self.base}/api2/auth-token/",
            data={"username": username, "password": password},
            timeout=15,
        )
        r.raise_for_status()
        self.session.headers.update({"Authorization": f"Token {r.json()['token']}"})
        log.info("Login successful.")

    def find_library(self, name):
        for endpoint in ("/api2/beshared-repos/", "/api2/repos/"):
            r = self.session.get(f"{self.base}{endpoint}", timeout=15)
            r.raise_for_status()
            for lib in r.json():
                if lib.get("name", "").lower() == name.lower():
                    return lib
        return None

    def list_dir(self, repo_id, path="/"):
        r = self.session.get(
            f"{self.base}/api2/repos/{repo_id}/dir/",
            params={"p": path}, timeout=15,
        )
        r.raise_for_status()
        return r.json()

    def download_file(self, repo_id, path):
        r = self.session.get(
            f"{self.base}/api2/repos/{repo_id}/file/",
            params={"p": path, "reuse": 1}, timeout=15,
        )
        r.raise_for_status()
        dl = self.session.get(r.json(), timeout=60)
        dl.raise_for_status()
        return dl.content

    def list_all_files(self, repo_id, path="/"):
        files = []
        try:
            entries = self.list_dir(repo_id, path)
        except Exception as e:
            log.warning(f"Could not list {path}: {e}")
            return files
        for entry in entries:
            if entry["type"] == "file":
                entry["full_path"] = path.rstrip("/") + "/" + entry["name"]
                files.append(entry)
            elif entry["type"] == "dir":
                files.extend(self.list_all_files(repo_id, path.rstrip("/") + "/" + entry["name"]))
        return files

def extract_grades_from_excel(data: bytes, student_number: str) -> dict:
    grades = {}
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        student_col = student_row_idx = None
        for r_idx, row in enumerate(rows):
            for c_idx, cell in enumerate(row):
                if cell is not None and str(cell).strip() == str(student_number).strip():
                    student_col, student_row_idx = c_idx, r_idx
                    break
            if student_col is not None:
                break
        if student_row_idx is None:
            continue
        header_row = None
        for h in range(max(0, student_row_idx - 5), student_row_idx):
            if sum(1 for c in rows[h] if c is not None) > 3:
                header_row = rows[h]
                break
        for c_idx, value in enumerate(rows[student_row_idx]):
            if value is None or c_idx == student_col:
                continue
            label = f"Col {c_idx+1}"
            if header_row and c_idx < len(header_row) and header_row[c_idx]:
                label = str(header_row[c_idx]).strip()
            grades[f"{sheet_name} — {label}"] = str(value)
    return grades


def extract_grades_from_pdf(data: bytes, student_number: str) -> dict:
    grades = {}
    with pdfplumber.open(BytesIO(data)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            # Try structured table extraction first
            tables = page.extract_tables()
            for t_idx, table in enumerate(tables):
                if not table:
                    continue
                header = table[0]
                for row in table[1:]:
                    if row and any(str(cell).strip() == str(student_number).strip() for cell in row if cell):
                        for c_idx, cell in enumerate(row):
                            if cell and str(cell).strip() != str(student_number).strip():
                                label = (
                                    str(header[c_idx]).strip()
                                    if c_idx < len(header) and header[c_idx]
                                    else f"Col {c_idx+1}"
                                )
                                grades[f"P{page_num} — {label}"] = str(cell).strip()
            # Fallback: raw text scan
            if not grades:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    if str(student_number) in line:
                        # Decimal-first so 15.5 is captured whole, not truncated to 15
                        numbers = re.findall(r"\b(\d{1,2}[.,]\d{1,2}|\d{1,2})\b", line)
                        grade_numbers = [
                            n for n in numbers
                            if 0 <= float(n.replace(",", ".")) <= 20
                            and n != str(student_number)
                        ]
                        if grade_numbers:
                            grades[f"P{page_num} — Grade"] = ", ".join(grade_numbers)
    return grades


def extract_grades_from_csv(data: bytes, student_number: str) -> dict:
    grades = {}
    reader = csv.DictReader(data.decode("utf-8", errors="replace").splitlines())
    for row in reader:
        if any(str(v).strip() == str(student_number).strip() for v in row.values()):
            for k, v in row.items():
                if v and str(v).strip() != str(student_number).strip():
                    grades[k] = v
            break
    return grades


def extract_grades(filename: str, data: bytes, student_number: str) -> dict:
    ext = Path(filename).suffix.lower()
    try:
        if ext in (".xlsx", ".xls", ".ods"):
            return extract_grades_from_excel(data, student_number)
        elif ext == ".pdf":
            return extract_grades_from_pdf(data, student_number)
        elif ext == ".csv":
            return extract_grades_from_csv(data, student_number)
        else:
            log.info(f"Skipping unsupported file type: {filename}")
            return {}
    except Exception as e:
        log.error(f"Error parsing {filename}: {e}")
        return {}

def send_notification(subject: str, filename: str, grades: dict):
    title = f"New grades: {subject}"
    lines = [f"File: {filename}", f"Subject: {subject}", ""]
    for label, value in grades.items():
        lines.append(f"  • {label}: {value}")
    body = "\n".join(lines)

    notified = False
    try:
        from plyer import notification
        notification.notify(title=title, message=body, app_name="Seafile Grade Monitor", timeout=10)
        notified = True
    except Exception:
        pass
    if not notified:
        try:
            from win10toast import ToastNotifier
            ToastNotifier().show_toast(title, body, duration=8, threaded=True)
        except Exception:
            log.warning("Could not send Windows notification. Install plyer: pip install plyer")

    border = "=" * 60
    print(f"\n{border}")
    print(f"  NEW GRADE: {title}")
    print(border)
    print(body)
    print(f"{border}\n")



def load_state() -> dict:
    if os.path.exists(CONFIG["state_file"]):
        with open(CONFIG["state_file"], "r") as f:
            return json.load(f)
    return {"seen_files": {}, "first_run": True}


def save_state(state: dict):
    with open(CONFIG["state_file"], "w") as f:
        json.dump(state, f, indent=2)


def file_hash(entry: dict) -> str:
    key = f"{entry.get('full_path','')}-{entry.get('mtime','')}-{entry.get('size','')}"
    return hashlib.md5(key.encode()).hexdigest()



def main():
    print("=" * 60)
    print("  Seafile Grade Monitor")
    print("=" * 60)

    if "your.email" in CONFIG["username"]:
        print("\n  Please edit the CONFIG section at the top of the script first!\n")
        sys.exit(1)

    state = load_state()
    seen = state.get("seen_files", {})
    first_run = state.get("first_run", True)

    client = SeafileClient(CONFIG["seafile_url"], CONFIG["username"], CONFIG["password"])

    repo = client.find_library(CONFIG["library_name"])
    if not repo:
        log.error(f"Library '{CONFIG['library_name']}' not found. Check library_name in CONFIG.")
        sys.exit(1)
    log.info(f"Found library: {repo['name']}")

    files = client.list_all_files(repo["id"], CONFIG["results_path"])

    if first_run:
        for entry in files:
            seen[file_hash(entry)] = {"path": entry["full_path"], "skipped_on_first_run": True}
        state["seen_files"] = seen
        state["first_run"] = False
        save_state(state)
        print("\n  First run complete — existing files recorded, no notifications sent.")
        print("  Run the script again after new grades are uploaded.\n")
        return
    
    new_count = 0
    for entry in files:
        fhash = file_hash(entry)
        if fhash in seen:
            continue

        filename = entry["name"]
        full_path = entry["full_path"]
        log.info(f"New file detected: {full_path}")

        try:
            data = client.download_file(repo["id"], full_path)
        except Exception as e:
            log.error(f"Could not download {full_path}: {e}")
            seen[fhash] = {"path": full_path, "processed": False}
            continue

        grades = extract_grades(filename, data, CONFIG["student_number"])

        if grades:
            parts = full_path.strip("/").split("/")
            subject = parts[-2] if len(parts) >= 2 else filename
            send_notification(subject, filename, grades)
            new_count += 1
        else:
            log.info(f"Student {CONFIG['student_number']} not found in {filename}")

        seen[fhash] = {
            "path": full_path,
            "processed": True,
            "grades_found": bool(grades),
            "checked_at": datetime.now().isoformat(),
        }

    state["seen_files"] = seen
    save_state(state)

    if new_count == 0:
        print("\n  No new grades since last check.\n")
    else:
        print(f"\n  Done — {new_count} new grade file(s) found.\n")


if __name__ == "__main__":
    main()